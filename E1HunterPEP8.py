# Osmar Fernando García González
# Raymundo Chong-Sing García Benavides

from pyhunter import PyHunter
from openpyxl import Workbook
import getpass


def Busqueda(organizacion):
    # Cantidad de resultados esperados de la búsqueda
    # El límite MENSUAL de Hunter es 50, cuidado!
    resultado = hunter.domain_search(
        company=organizacion,
        limit=1,
        emails_type='personal')
    return resultado


def GuardarInformacion(datosEncontrados, organizacion):
    libro = Workbook()
    hoja = libro.create_sheet(organizacion)
    libro.save("Hunter" + organizacion + ".xlsx")
    # Agrega el codigo necesario para guardar en formato tabla
    # dentro del libro de Excel, información que consideres relevante
    # de lo obtenido en la búsqueda.
    libro.active = 1
    hoja = libro.active
    hoja.append(('Pagina Web', 'Correo Electronico', 'Tipo de Correo'))
    count = 2
    correos = datosEncontrados['emails']
    for x in correos:
        hoja.cell(count, 1, organizacion)
        hoja.cell(count, 2, x['value'])
        hoja.cell(count, 3, x['type'])
        count += 1
    libro.save("Hunter" + organizacion + ".xlsx")


print("Script para buscar información")
APIKey = getpass.getpass("Ingresa tu API key: ")
hunter = PyHunter(APIKey)
orga = input("Dominio a investigar: ")
datosEncontrados = Busqueda(orga)
if datosEncontrados is None:
    exit()
else:
    print(datosEncontrados)
    print(type(datosEncontrados))
    GuardarInformacion(datosEncontrados, orga)
